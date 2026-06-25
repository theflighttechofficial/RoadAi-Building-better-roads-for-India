"""
excel_exporter.py — Road-AI Excel Report Generator
"""
from __future__ import annotations
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict

log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BG_DARK="FF0D1117"; BG_MED="FF161B22"; BG_LIGHT="FF1C2235"
ACCENT="FF10D98A"; TEXT_WHITE="FFE2E8F0"; TEXT_DIM="FF7C8299"
GOOD="FF10D98A"; MOD="FFF5A623"; POOR="FFF04444"; CRIT="FFA855F7"
BLUE="FF89B4FA"; TEAL="FF94E2D5"; BORDER_COL="FF21262D"

def _fill(h):  return PatternFill("solid", fgColor=h)
def _font(bold=False, color=TEXT_WHITE, size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center")

def score_color(s):
    if s >= 80: return GOOD
    if s >= 60: return MOD
    if s >= 40: return POOR
    return CRIT

def score_tier(s):
    if s >= 80: return "Good"
    if s >= 60: return "Moderate"
    if s >= 40: return "Poor"
    return "Critical"

def _hdr(ws, row, col, text, color=ACCENT):
    c = ws.cell(row, col, text)
    c.fill = _fill(BG_DARK)
    c.font = _font(bold=True, color=color[2:])
    c.alignment = _center()
    c.border = _border()
    return c

def _cell(ws, row, col, val, bg=BG_MED, color=TEXT_WHITE, bold=False):
    c = ws.cell(row, col, val)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=color[2:] if len(color) == 8 else color)
    c.alignment = _center()
    c.border = _border()
    return c


class ExcelExporter:
    def __init__(self):
        if not OPENPYXL_OK:
            raise RuntimeError("openpyxl not installed. pip install openpyxl")

    def export(self, frames: List[Dict], summary: Dict,
               output_path: str = "output/report.xlsx") -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._summary(wb, frames, summary)
        self._frames(wb, frames)
        self._classes(wb, frames)
        self._flagged(wb, frames)
        self._conditions(wb, frames)
        self._gps(wb, frames)
        self._measurements(wb, frames)
        wb.save(output_path)
        log.info(f"Excel saved → {output_path}")
        return output_path

    # ── 1. Summary ────────────────────────────────────────────────
    def _summary(self, wb, frames, summary):
        ws = wb.create_sheet("📊 Summary")
        ws.sheet_view.showGridLines = False
        for col_idx, w in enumerate([26,22,4,4,14,14,14,14,14,14], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Title
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = "ROAD-AI  —  ROAD DAMAGE ASSESSMENT REPORT"
        c.fill = _fill(BG_DARK)
        c.font = Font(name="Arial", bold=True, color=ACCENT[2:], size=15)
        c.alignment = _center()
        ws.row_dimensions[1].height = 44

        ws.merge_cells("A2:H2")
        c = ws["A2"]
        sid = summary.get("session_id","—")[:8].upper()
        c.value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}   |   Session: {sid}"
        c.fill = _fill(BG_MED)
        c.font = _font(color=TEXT_DIM[2:], size=8)
        c.alignment = _center()
        ws.row_dimensions[2].height = 18

        scores  = [f.get("health_score", 100) for f in frames]
        n       = len(frames)
        avg     = sum(scores)/n if scores else 100
        mn      = min(scores) if scores else 100
        tot_det = sum(f.get("n_detections", 0) for f in frames)
        tot_cst = sum(f.get("cost_inr", 0) for f in frames)
        flagged = sum(1 for s in scores if s < 70)
        cond_counts = {}
        for f in frames:
            c_ = (f.get("condition") or "normal").lower()
            cond_counts[c_] = cond_counts.get(c_, 0) + 1

        # KPI header
        ws.merge_cells("A4:B4")
        c = ws["A4"]
        c.value = "KEY PERFORMANCE INDICATORS"
        c.fill = _fill(BG_LIGHT)
        c.font = _font(bold=True, color=ACCENT[2:], size=9)
        c.alignment = _center()
        ws.row_dimensions[4].height = 20

        kpis = [
            ("Avg Health Score",   f"{avg:.1f} / 100",     score_color(avg)),
            ("Min Health Score",   f"{mn:.1f} / 100",      score_color(mn)),
            ("Road Condition",     score_tier(avg),         score_color(avg)),
            ("Total Frames",       f"{n:,}",                BLUE),
            ("Frames w/ Damage",   f"{sum(1 for f in frames if f.get('n_detections',0)>0):,}", MOD),
            ("Total Detections",   f"{tot_det:,}",          MOD),
            ("Flagged (Score<70)", f"{flagged:,}",          POOR if flagged>5 else MOD),
            ("Est. Repair Cost",   f"\u20b9{tot_cst:,.0f}", POOR),
            ("Repair Urgency",     "Immediate" if avg<40 else "Within 7 days" if avg<60 else "Within 30 days", MOD),
            ("Dominant Condition", max(cond_counts, key=cond_counts.get) if cond_counts else "Normal", TEAL),
        ]

        for i, (label, value, color) in enumerate(kpis):
            r = 5 + i
            ws.row_dimensions[r].height = 26
            ca = ws.cell(r, 1, label)
            ca.fill = _fill(BG_MED); ca.font = _font(color=TEXT_DIM[2:])
            ca.alignment = _left(); ca.border = _border()
            cb = ws.cell(r, 2, value)
            cb.fill = _fill(BG_DARK)
            cb.font = Font(name="Arial", bold=True, color=color[2:], size=11)
            cb.alignment = _center(); cb.border = _border()

        # Chart data in cols E–H (separate from KPIs)
        _hdr(ws, 4, 5, "Frame", BLUE)
        _hdr(ws, 4, 6, "Score", GOOD)
        _hdr(ws, 4, 7, "Cost", MOD)
        _hdr(ws, 4, 8, "Detections", POOR)
        sample = frames[::max(1, n//80)]
        for i, f in enumerate(sample):
            r = 5 + i
            ws.cell(r, 5, f.get("frame_index", i))
            ws.cell(r, 6, round(f.get("health_score", 100), 1))
            ws.cell(r, 7, round(f.get("cost_inr", 0), 0))
            ws.cell(r, 8, f.get("n_detections", 0))
        npts = len(sample)

        # Score trend line chart
        ch1 = LineChart()
        ch1.title = "Health Score Trend"; ch1.style = 10
        ch1.y_axis.title = "Score /100"; ch1.x_axis.title = "Frame"
        ch1.y_axis.scaling.min = 0; ch1.y_axis.scaling.max = 100
        ch1.width = 22; ch1.height = 11
        d1 = Reference(ws, min_col=6, min_row=4, max_row=4+npts)
        c1 = Reference(ws, min_col=5, min_row=5, max_row=4+npts)
        ch1.add_data(d1, titles_from_data=True); ch1.set_categories(c1)
        ch1.series[0].graphicalProperties.line.solidFill = "10D98A"
        ch1.series[0].graphicalProperties.line.width = 18000
        ws.add_chart(ch1, "A16")

        # Detection bar chart
        ch2 = BarChart()
        ch2.title = "Detections per Frame"; ch2.style = 10
        ch2.y_axis.title = "Detections"; ch2.width = 22; ch2.height = 10
        d2 = Reference(ws, min_col=8, min_row=4, max_row=4+npts)
        ch2.add_data(d2, titles_from_data=True); ch2.set_categories(c1)
        ch2.series[0].graphicalProperties.solidFill = "F04444"
        ws.add_chart(ch2, "A32")

    # ── 2. Frame Data ─────────────────────────────────────────────
    def _frames(self, wb, frames):
        ws = wb.create_sheet("🎞 Frame Data")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        headers = ["Frame #","Time (s)","Detections","Health Score",
                   "Condition","Dominant Class","Cost (\u20b9)",
                   "Width cm","Depth cm","Severity","GPS Lat","GPS Lon","Enhanced"]
        widths  = [10,10,12,14,12,22,14,10,10,12,14,14,10]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(col)].width = w
            _hdr(ws, 1, col, h)
        ws.row_dimensions[1].height = 22

        for row, f in enumerate(frames, 2):
            score = f.get("health_score", 100)
            cond  = (f.get("condition") or "normal").upper()
            bg    = BG_MED if row%2==0 else BG_DARK
            vals  = [
                f.get("frame_index", row-2),
                round(f.get("timestamp_sec", 0), 2),
                f.get("n_detections", 0),
                round(score, 1),
                cond,
                f.get("dominant_class","") or "—",
                round(f.get("cost_inr", 0)),
                round(f.get("max_width_cm", 0) or 0, 1),
                round(f.get("max_depth_cm", 0) or 0, 1),
                round(f.get("max_severity",  0) or 0, 1),
                f.get("gps_lat") or "—",
                f.get("gps_lon") or "—",
                "Yes" if f.get("enhanced") else "No",
            ]
            for col, val in enumerate(vals, 1):
                col_c = score_color(score)[2:] if col==4 else TEXT_WHITE[2:]
                _cell(ws, row, col, val, bg=bg, color=col_c)
            ws.row_dimensions[row].height = 15

        n = len(frames)
        ws.conditional_formatting.add(f"D2:D{n+1}", ColorScaleRule(
            start_type="num", start_value=0,  start_color="F04444",
            mid_type="num",   mid_value=60,   mid_color="F5A623",
            end_type="num",   end_value=100,  end_color="10D98A"))
        ws.conditional_formatting.add(f"G2:G{n+1}",
            DataBarRule(start_type="min", end_type="max", color="F5A623"))

    # ── 3. Damage Classes ─────────────────────────────────────────
    def _classes(self, wb, frames):
        ws = wb.create_sheet("🔍 Damage Classes")
        ws.sheet_view.showGridLines = False
        for ci_, w in enumerate([26,12,16,18,14], 1):
            ws.column_dimensions[get_column_letter(ci_)].width = w
        for col, h in enumerate(["Damage Class","Occurrences","Total Cost (\u20b9)","Avg Cost/Det (\u20b9)","% of Total"], 1):
            _hdr(ws, 1, col, h)
        ws.row_dimensions[1].height = 22

        # Count from full detections array first, fallback to dominant_class
        class_data: Dict = {}
        for f in frames:
            dets = f.get("detections", [])
            if dets:
                for d in dets:
                    cls = d.get("class_name","") or ""
                    cost = d.get("cost_inr", 0) or 0
                    if cls:
                        if cls not in class_data:
                            class_data[cls] = {"count":0,"cost":0.0}
                        class_data[cls]["count"] += 1
                        class_data[cls]["cost"]  += cost
            else:
                cls  = f.get("dominant_class","") or ""
                cost = f.get("cost_inr", 0) or 0
                ndet = f.get("n_detections", 0) or 0
                if cls and ndet > 0:
                    if cls not in class_data:
                        class_data[cls] = {"count":0,"cost":0.0}
                    class_data[cls]["count"] += ndet
                    class_data[cls]["cost"]  += cost

        total_count = sum(d["count"] for d in class_data.values()) or 1
        colors_list = [GOOD, MOD, POOR, CRIT, BLUE, TEAL]
        sorted_cls  = sorted(class_data.items(), key=lambda x:-x[1]["cost"])

        for row,(cls,d) in enumerate(sorted_cls, 2):
            col_c = colors_list[(row-2) % len(colors_list)]
            avg_c = d["cost"]/d["count"] if d["count"] else 0
            pct   = round(d["count"]/total_count*100, 1)
            bg    = BG_MED if row%2==0 else BG_DARK
            for col,val in enumerate([cls, d["count"], round(d["cost"]), round(avg_c), f"{pct}%"], 1):
                _cell(ws, row, col, val, bg=bg, color=col_c[2:] if col==1 else TEXT_WHITE[2:], bold=(col==1))
            ws.row_dimensions[row].height = 20

        n = len(sorted_cls)
        if n > 0:
            tr = n+2
            ws.row_dimensions[tr].height = 22
            _cell(ws, tr, 1, "TOTAL",              bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 2, f"=SUM(B2:B{tr-1})", bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 3, f"=SUM(C2:C{tr-1})", bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 4, f"=IFERROR(C{tr}/B{tr},0)", bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 5, "100%",               bg=BG_LIGHT, color=ACCENT[2:], bold=True)

            # Pie chart — damage distribution
            pie = PieChart()
            pie.title = "Damage Type Distribution"; pie.style = 10
            pie.width = 16; pie.height = 12
            pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=1+n), titles_from_data=True)
            pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1+n))
            ws.add_chart(pie, "G2")

            # Bar chart — cost by class
            bar = BarChart()
            bar.type = "bar"; bar.title = "Total Cost by Damage Type (\u20b9)"; bar.style = 10
            bar.width = 16; bar.height = 12
            bar.add_data(Reference(ws, min_col=3, min_row=1, max_row=1+n), titles_from_data=True)
            bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1+n))
            bar.series[0].graphicalProperties.solidFill = "F04444"
            ws.add_chart(bar, "G18")

    # ── 4. Flagged Frames ─────────────────────────────────────────
    def _flagged(self, wb, frames):
        ws = wb.create_sheet("🚨 Flagged Frames")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        flagged = sorted([f for f in frames if f.get("health_score",100) < 70],
                         key=lambda x: x.get("health_score",100))
        headers = ["Frame #","Time (s)","Health Score","Severity","Detections",
                   "Dominant Class","Cost (\u20b9)","Width cm","Depth cm","GPS Lat","GPS Lon"]
        widths  = [10,10,14,12,12,24,14,10,10,14,14]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(col)].width = w
            _hdr(ws, 1, col, h, color=POOR)
        ws.row_dimensions[1].height = 22

        if not flagged:
            ws.merge_cells("A2:K2")
            c = ws["A2"]
            c.value = "✅  No frames with Health Score < 70 — Road is in good condition"
            c.fill = _fill(BG_MED); c.font = _font(color=GOOD[2:], bold=True)
            c.alignment = _center()
            return

        for row, f in enumerate(flagged, 2):
            score = f.get("health_score", 0)
            sev   = "Critical" if score < 40 else "Poor"
            bg    = "FF1A0505" if row%2==0 else BG_DARK
            vals  = [f.get("frame_index",0), round(f.get("timestamp_sec",0),2),
                     round(score,1), sev, f.get("n_detections",0),
                     f.get("dominant_class","—") or "—", round(f.get("cost_inr",0)),
                     round(f.get("max_width_cm",0) or 0,1),
                     round(f.get("max_depth_cm",0) or 0,1),
                     f.get("gps_lat") or "—", f.get("gps_lon") or "—"]
            for col,val in enumerate(vals,1):
                _cell(ws, row, col, val, bg=bg,
                      color=score_color(score)[2:] if col==3 else TEXT_WHITE[2:])
            ws.row_dimensions[row].height = 16

    # ── 5. Conditions ─────────────────────────────────────────────
    def _conditions(self, wb, frames):
        ws = wb.create_sheet("🌦 Conditions")
        ws.sheet_view.showGridLines = False
        for ci__, w in enumerate([22,14,14,18], 1):
            ws.column_dimensions[get_column_letter(ci__)].width = w
        for col,h in enumerate(["Condition","Frame Count","% of Total","Avg Health Score"],1):
            _hdr(ws, 1, col, h, color=TEAL)
        ws.row_dimensions[1].height = 22

        cond_data: Dict = {}
        for f in frames:
            c_ = (f.get("condition") or "normal").lower()
            s  = f.get("health_score", 100)
            if c_ not in cond_data:
                cond_data[c_] = {"count":0,"scores":[]}
            cond_data[c_]["count"] += 1
            cond_data[c_]["scores"].append(s)

        icons  = {"night":"🌙","rain":"🌧","fog":"🌫","glare":"☀","normal":"✅","clear":"✅"}
        cols   = {"night":BLUE,"rain":TEAL,"fog":TEXT_DIM,"glare":MOD,"normal":GOOD,"clear":GOOD}
        total  = len(frames) or 1

        for row,(cond,d) in enumerate(sorted(cond_data.items(), key=lambda x:-x[1]["count"]),2):
            avg_s = sum(d["scores"])/len(d["scores"]) if d["scores"] else 100
            pct   = round(d["count"]/total*100, 1)
            icon  = icons.get(cond, "❓")
            col_c = cols.get(cond, TEXT_WHITE)
            bg    = BG_MED if row%2==0 else BG_DARK
            for col,val in enumerate([f"{icon}  {cond.title()}", d["count"], f"{pct}%", round(avg_s,1)],1):
                _cell(ws, row, col, val, bg=bg, color=col_c[2:] if col==1 else TEXT_WHITE[2:], bold=(col==1))
            ws.row_dimensions[row].height = 22

        n = len(cond_data)
        if n > 0:
            tr = n+2
            _cell(ws, tr, 1, "TOTAL",                    bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 2, f"=SUM(B2:B{tr-1})",       bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 3, "100%",                     bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            _cell(ws, tr, 4, f"=IFERROR(AVERAGE(D2:D{tr-1}),100)", bg=BG_LIGHT, color=ACCENT[2:], bold=True)
            ws.row_dimensions[tr].height = 22

            bar = BarChart()
            bar.type = "col"; bar.title = "Frames by Condition"; bar.style = 10
            bar.y_axis.title = "Frames"; bar.width = 18; bar.height = 12
            bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=1+n), titles_from_data=True)
            bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1+n))
            bar.series[0].graphicalProperties.solidFill = "89B4FA"
            ws.add_chart(bar, "F2")

    # ── 6. GPS Track ──────────────────────────────────────────────
    def _gps(self, wb, frames):
        gps_frames = [f for f in frames if f.get("gps_lat")]
        if not gps_frames: return
        ws = wb.create_sheet("📍 GPS Track")
        ws.sheet_view.showGridLines = False
        headers = ["Frame #","Latitude","Longitude","Speed (km/h)","Health Score","Cost (\u20b9)","Condition"]
        widths  = [10,16,16,14,14,14,14]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(col)].width = w
            _hdr(ws, 1, col, h, color=BLUE)
        ws.row_dimensions[1].height = 22
        for row,f in enumerate(gps_frames,2):
            score = f.get("health_score",100)
            bg    = BG_MED if row%2==0 else BG_DARK
            vals  = [f.get("frame_index",0), round(f.get("gps_lat",0),6),
                     round(f.get("gps_lon",0),6), round(f.get("speed_kmh",0) or 25,1),
                     round(score,1), round(f.get("cost_inr",0)),
                     (f.get("condition") or "normal").title()]
            for col,val in enumerate(vals,1):
                _cell(ws, row, col, val, bg=bg,
                      color=score_color(score)[2:] if col==5 else TEXT_WHITE[2:])
            ws.row_dimensions[row].height = 15
        n = len(gps_frames)
        ws.conditional_formatting.add(f"E2:E{n+1}", ColorScaleRule(
            start_type="num", start_value=0,  start_color="F04444",
            mid_type="num",   mid_value=60,   mid_color="F5A623",
            end_type="num",   end_value=100,  end_color="10D98A"))

    # ── 7. Measurements ───────────────────────────────────────────
    def _measurements(self, wb, frames):
        has_data = any(f.get("max_width_cm",0) or f.get("max_depth_cm",0) for f in frames)
        if not has_data: return
        ws = wb.create_sheet("📐 Measurements")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        headers = ["Frame #","Dominant Class","Width (cm)","Length (cm)",
                   "Depth (cm)","Area (cm\u00b2)","Severity /100","Cost (\u20b9)","Uncertain"]
        widths  = [10,22,12,12,12,12,14,14,10]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            ws.column_dimensions[get_column_letter(col)].width = w
            _hdr(ws, 1, col, h, color=TEAL)
        ws.row_dimensions[1].height = 22
        row = 2
        for f in frames:
            dets = f.get("detections",[])
            if dets:
                for d in dets:
                    sev = d.get("severity",0) or 0
                    bg  = BG_MED if row%2==0 else BG_DARK
                    for col,val in enumerate([
                        f.get("frame_index",0), d.get("class_name","—"),
                        round(d.get("width_cm",0) or 0,1), round(d.get("length_cm",0) or 0,1),
                        round(d.get("depth_cm",0) or 0,1), round(d.get("area_cm2",0) or 0,1),
                        round(sev,1), round(d.get("cost_inr",0) or 0),
                        "⚠" if d.get("uncertain") else "✅"],1):
                        _cell(ws, row, col, val, bg=bg,
                              color=score_color(100-sev)[2:] if col==7 else TEXT_WHITE[2:])
                    ws.row_dimensions[row].height = 15; row += 1
            elif f.get("max_width_cm") or f.get("max_severity"):
                sev = f.get("max_severity",0) or 0
                bg  = BG_MED if row%2==0 else BG_DARK
                for col,val in enumerate([
                    f.get("frame_index",0), f.get("dominant_class","—") or "—",
                    round(f.get("max_width_cm",0) or 0,1), 0,
                    round(f.get("max_depth_cm",0) or 0,1), 0,
                    round(sev,1), round(f.get("cost_inr",0)), "—"],1):
                    _cell(ws, row, col, val, bg=bg,
                          color=score_color(100-sev)[2:] if col==7 else TEXT_WHITE[2:])
                ws.row_dimensions[row].height = 15; row += 1